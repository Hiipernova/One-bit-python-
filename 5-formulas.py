from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("data/barchart.xlsx")
sheet = wb["Relatorio"]

min_column = sheet.min_column
max_column = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

for i in range(min_column + 1, max_column + 1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"

wb.save("test.xlsx")

